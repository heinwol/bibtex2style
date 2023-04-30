# import os
import argparse
import subprocess
import pathlib

import fitz

from toolz import *

import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from lenses import lens
import glom

# print(os.getcwd()) 
# print(subprocess.run('ls').stdout) 
# it = pathlib.Path(__file__) 
# print(it.resolve().parent)
 
def flags_decomposer(flags):
    """Make font flags human readable."""
    l = []
    if flags & 2 ** 0:
        l.append("superscript")
    if flags & 2 ** 1:
        l.append("italic")
    if flags & 2 ** 2:
        l.append("serifed")
    else:
        l.append("sans")
    if flags & 2 ** 3:
        l.append("monospaced")
    else:
        l.append("proportional")
    if flags & 2 ** 4:
        l.append("bold")
    # for now, we're interested only in bold and italic
    l = list(filter(lambda x: x in ["bold", "italic"], l))
    return l

get_text_and_formatting_from_span = flip(glom.glom)({
    'text': 'text',
    'attributes': glom.Invoke(flags_decomposer).specs('flags')
})

def text_dict_to_openpyxl_text(d):
    if d['attributes'] == []:
        return d['text']
    else:
        return TextBlock(InlineFont(
            b=('bold' in d['attributes']),
            i=('italic' in d['attributes'])
        ),
        d['text'])

def parse_pdf(doc):
    # we have only one page
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    # we are not interested in title
    blocks = blocks[1:]

    bib_list = thread_first(
        blocks,
        lens.Each()['lines'].F(lens.Each()['spans'].Each().collect()).collect(),
        lens.Each().Each().modify(get_text_and_formatting_from_span),
        lens.Each().modify(lambda x: (x[0], x[2:])),
        lens.Each()[0].modify(text_dict_to_openpyxl_text),
        lens.Each()[1].Each().modify(text_dict_to_openpyxl_text),
        lens.Each()[1].modify(lambda x: CellRichText(*x)),
    )

doc = fitz.open("res.pdf")

wb = openpyxl.Workbook()
ws = wb.active

for row in bib_list:
    ws.append(row)
    
max_first_col_width = max(len(cell.value or "") for cell in ws['A'])
ws.column_dimensions['A'].width = max_first_col_width*1.3
ws.column_dimensions['B'].width = 85
for i in range(1, ws.max_row + 1):
    ws.row_dimensions[i].height = 65

first_col_alignment = openpyxl.styles.Alignment(
    horizontal='center',
    vertical='center'
)
ws.column_dimensions['A'].alignment = first_col_alignment

second_col_alignment = openpyxl.styles.Alignment(
    wrapText = True
)
ws.column_dimensions['B'].alignment = second_col_alignment

ws.sheet_view.zoomScale = 100

wb.save('temp.xlsx')