import os
import argparse
import subprocess
import pathlib
import shutil
from time import sleep

from typing import List, Tuple, Dict

# provided by PyMuPDF
import fitz

from toolz.curried import *

import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from lenses import lens
import glom

def flags_decomposer(flags) -> List[str]:
    """Make font flags human readable."""
    l = []
    if flags & 2 ** 0:
        l.append("superscript")
    if flags & 2 ** 1:
        l.append("italic")
    if flags & 2 ** 2:
        l.append("serifed")
    else:
        l.append("sans")
    if flags & 2 ** 3:
        l.append("monospaced")
    else:
        l.append("proportional")
    if flags & 2 ** 4:
        l.append("bold")
    # for now, we're interested only in bold and italic
    l = list(filter(lambda x: x in ["bold", "italic"], l))
    return l

get_text_and_formatting_from_span = flip(glom.glom)({
    'text': 'text',
    'attributes': glom.Invoke(flags_decomposer).specs('flags')
})

def text_dict_to_openpyxl_text(d) -> TextBlock:
    if d['attributes'] == []:
        return d['text']
    else:
        return TextBlock(InlineFont(
            b=('bold' in d['attributes']),
            i=('italic' in d['attributes'])
        ),
        d['text'])

def wait_for_file_to_be_available(fname: str) -> None:
    """
        at least on windows and with miktex, the latexmk
        command exits prior to the pdf file being actually closed.
        We have to wait then 
    """
    for i in range(7):
        try:
            with open(fname, 'rb'):
                pass
            return
        except PermissionError:
            print('trying to access .pdf file...')
            sleep(0.3)
    with open(fname, 'rb'):
        pass

def parse_pdf(doc: fitz.Document) -> List[Tuple[TextBlock, CellRichText]]:
    # we have only one page
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    # we are not interested in title
    blocks = blocks[1:]

    bib_list = thread_first(
        blocks,
        lens.Each()['lines'].F(lens.Each()['spans'].Each().collect()).collect(),
        lens.Each().Each().modify(get_text_and_formatting_from_span),
        lens.Each().modify(lambda x: (x[0], x[2:])),
        lens.Each()[0].modify(text_dict_to_openpyxl_text),
        lens.Each()[1].Each().modify(text_dict_to_openpyxl_text),
        lens.Each()[1].modify(lambda x: CellRichText(*x)),
    )
    return bib_list

def bib_list_to_spreadsheet(bib_list: List[Tuple[TextBlock, CellRichText]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active

    for row in bib_list:
        ws.append(row)
    
    max_first_col_width = max(len(cell.value or "") for cell in ws['A'])
    ws.column_dimensions['A'].width = max_first_col_width*1.3
    ws.column_dimensions['B'].width = 85
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 65

    first_col_alignment = openpyxl.styles.Alignment(
        horizontal='center',
        vertical='center'
    )
    ws.column_dimensions['A'].alignment = first_col_alignment

    second_col_alignment = openpyxl.styles.Alignment(
        wrapText = True
    )
    ws.column_dimensions['B'].alignment = second_col_alignment

    ws.sheet_view.zoomScale = 100
    return wb

def main():

    parser = argparse.ArgumentParser(description='Process a .bib file to get .xlsx with styling')
    parser.add_argument('bib_file',
                        type=str,
                        help='.bib bibliography file to convert')
    parser.add_argument('xlsx_file',
                        nargs='?',
                        default='bibliography.xlsx',
                        help='(optional) path to desired .xlsx file')
    args = parser.parse_args()

    bib_file = pathlib.Path(args.bib_file).resolve()
    xlsx_file = pathlib.Path(args.xlsx_file).resolve()
    current_dir = pathlib.Path(os.getcwd()).resolve()
    project_dir = pathlib.Path(__file__).resolve().parent
    tempdir = current_dir.joinpath('.bibtex2style')

    if tempdir.exists():
        shutil.rmtree(tempdir)
    os.mkdir(tempdir)
    os.chdir(tempdir)
    shutil.copy(project_dir.joinpath('process_bib_file.tex'), tempdir)
    shutil.copyfile(bib_file, tempdir.joinpath('bib_file.bib'))

    latexcmd_res = subprocess.run(['latexmk', '-pdflatex=lualatex', '-pdf'])
    if latexcmd_res.returncode != 0:
        raise ChildProcessError('latex terminated with error, check logs\n')
    else:
        wait_for_file_to_be_available('process_bib_file.pdf')
        with fitz.open('process_bib_file.pdf') as doc:
            bib_list = parse_pdf(doc)
        wb = bib_list_to_spreadsheet(bib_list)
        
        wb.save('temp.xlsx')
        shutil.copyfile('temp.xlsx', xlsx_file)
        shutil.rmtree(tempdir)

if __name__ == '__main__':
    main()